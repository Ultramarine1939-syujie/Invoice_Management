"""CSV and Excel export builders."""

from __future__ import annotations

import csv
import io
from collections import defaultdict

from services.records import DONE_STATUS, FIELDNAMES_EXT, NUMERIC_FIELDS, amount_value


def build_csv(records: list[dict]) -> io.BytesIO:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=FIELDNAMES_EXT, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(records)

    csv_bytes = io.BytesIO(("\ufeff" + buf.getvalue()).encode("utf-8"))
    csv_bytes.seek(0)
    return csv_bytes


def build_excel(records: list[dict]) -> io.BytesIO:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("服务端未安装 openpyxl") from exc

    wb = openpyxl.Workbook()
    header_fill = PatternFill("solid", fgColor="1677FF")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin = Side(style="thin", color="D9D9D9")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws1 = wb.active
    ws1.title = "发票明细"
    ws1.append(FIELDNAMES_EXT)
    for col_idx, _field in enumerate(FIELDNAMES_EXT, 1):
        cell = ws1.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = cell_border

    for row_idx, rec in enumerate(records, 2):
        for col_idx, field in enumerate(FIELDNAMES_EXT, 1):
            val = rec.get(field, "")
            if field in NUMERIC_FIELDS and val:
                val = amount_value(val)
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.border = cell_border
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F5F8FF")

    for col_idx, field in enumerate(FIELDNAMES_EXT, 1):
        max_len = len(field)
        for row_idx in range(2, len(records) + 2):
            val = ws1.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws1.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)
    ws1.freeze_panes = "A2"

    ws2 = wb.create_sheet("费用汇总")
    total_amount = sum(amount_value(r.get("金额")) for r in records)
    total_tax = sum(amount_value(r.get("税额")) for r in records)
    total_all = sum(amount_value(r.get("价税合计")) for r in records)
    done_total = sum(amount_value(r.get("价税合计")) for r in records if r.get("报销状态") == DONE_STATUS)
    pending_total = total_all - done_total

    summary_data = [
        ("项目", "数值"),
        ("发票总数", len(records)),
        ("金额合计（元）", round(total_amount, 2)),
        ("税额合计（元）", round(total_tax, 2)),
        ("价税合计（元）", round(total_all, 2)),
        ("已报销（元）", round(done_total, 2)),
        ("待报销（元）", round(pending_total, 2)),
    ]
    for row_idx, (label, val) in enumerate(summary_data, 1):
        c1 = ws2.cell(row=row_idx, column=1, value=label)
        c2 = ws2.cell(row=row_idx, column=2, value=val)
        c1.border = cell_border
        c2.border = cell_border
        if row_idx == 1:
            c1.fill = header_fill
            c1.font = header_font
            c2.fill = header_fill
            c2.font = header_font
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 18

    ws3 = wb.create_sheet("销售方统计")
    seller_map = defaultdict(lambda: {"count": 0, "amount": 0.0})
    for record in records:
        name = record.get("销售方名称") or "未知"
        seller_map[name]["count"] += 1
        seller_map[name]["amount"] += amount_value(record.get("价税合计"))

    headers3 = ["销售方", "发票数量", "金额合计（元）"]
    for col, header in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = cell_border
    for row_idx, (name, data) in enumerate(sorted(seller_map.items(), key=lambda item: -item[1]["amount"]), 2):
        ws3.cell(row=row_idx, column=1, value=name).border = cell_border
        ws3.cell(row=row_idx, column=2, value=data["count"]).border = cell_border
        ws3.cell(row=row_idx, column=3, value=round(data["amount"], 2)).border = cell_border
    ws3.column_dimensions["A"].width = 40
    ws3.column_dimensions["B"].width = 14
    ws3.column_dimensions["C"].width = 18

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
